from flask import Flask, render_template, request, redirect, url_for, send_file
import sqlite3
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from datetime import datetime
from pathlib import Path

from werkzeug.utils import secure_filename

# Ensure this directory exists at startup
RECEIPTS_DIR = Path("static/receipts")
RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
DB = "expenses.db"

# --- helpers ---
def get_conn():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def init_db():
    with get_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS providers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            );
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                payment_ref TEXT DEFAULT 'NA',
                factura_ref TEXT 'NA',
                proveedor_id INTEGER NOT NULL,
                payment_type TEXT NOT NULL CHECK (payment_type IN ('na', 'cash','card','transfer','sinpe')),
                amount REAL NOT NULL,
                currency TEXT NOT NULL CHECK (currency IN ('CRC','USD')),
                details TEXT,
                delivered_email INTEGER NOT NULL DEFAULT 0,
                factura_aparte INTEGER NOT NULL DEFAULT 0,
                receipt_path TEXT '',
                FOREIGN KEY (proveedor_id) REFERENCES providers(id) ON DELETE RESTRICT
            );
            """
        )


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # --- handle adding a new expense ---
        delivered_email = 1 if "delivered_email" in request.form else 0
        factura_aparte = 1 if "factura_aparte" in request.form else 0

        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO expenses
                (date, payment_ref, factura_ref, proveedor_id, payment_type, amount, currency, details, delivered_email, factura_aparte, receipt_path)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    request.form["date"],
                    request.form.get("payment_ref", ""),
                    request.form.get("factura_ref", ""),
                    int(request.form["proveedor_id"]),
                    request.form["payment_type"],
                    float(request.form["amount"]),
                    request.form["currency"],
                    request.form.get("details", ""),
                    delivered_email,
                    factura_aparte,
                    request.form.get("receipt_path", "")
                ),
            )
        return redirect(url_for("index"))

    # --- GET request ---
    provider = request.args.get("provider")
    payment_type = request.args.get("payment_type")
    
    # Determine current month and default if no month filter
    current_month = datetime.now().strftime("%Y-%m")
    month = request.args.get("month") or current_month

    query = """
        SELECT e.id, e.date, e.proveedor_id, e.payment_ref, e.factura_ref, p.name as proveedor,
               e.payment_type, e.amount, e.currency, e.details,
               e.delivered_email, e.factura_aparte, e.receipt_path
        FROM expenses e
        JOIN providers p ON p.id = e.proveedor_id
        WHERE 1=1
    """
    params = []

    if month:
        query += " AND strftime('%Y-%m', e.date) = ?"
        params.append(month)
    if provider:
        query += " AND e.proveedor_id = ?"
        params.append(provider)
    if payment_type:
        query += " AND e.payment_type = ?"
        params.append(payment_type)

    query += " ORDER BY e.date DESC"

    with get_conn() as conn:
        expenses = conn.execute(query, params).fetchall()

    
    with sqlite3.connect(DB) as conn:
        conn.row_factory = sqlite3.Row
        expenses = conn.execute(query, params).fetchall()
        providers = conn.execute("SELECT * FROM providers").fetchall()

    stats = {
        "total_records": len(expenses),
        "total_crc": sum(e["amount"] for e in expenses if e["currency"] == "CRC"),
        "total_usd": sum(e["amount"] for e in expenses if e["currency"] == "USD"),
        "factura_aparte_count": sum(1 for e in expenses if e["factura_aparte"]),
        "factura_aparte_crc": sum(e["amount"] for e in expenses if e["factura_aparte"] and e["currency"] == "CRC") or 0,
        "factura_aparte_usd": sum(e["amount"] for e in expenses if e["factura_aparte"] and e["currency"] == "USD") or 0,
    }

    return render_template("index.html", expenses=expenses, providers=providers, current_month=current_month, stats=stats)


@app.route("/edit/<int:id>", methods=["POST"])
def edit(id):
    delivered_email = 1 if "delivered_email" in request.form else 0
    factura_aparte = 1 if "factura_aparte" in request.form else 0

    with get_conn() as conn:
        # Get current receipt path
        e = conn.execute("SELECT receipt_path FROM expenses WHERE id=?", (id,)).fetchone()
        current_receipt = e["receipt_path"] if e else ""

    # Handle file upload
    file = request.files.get("receipt_path")  # name in your <input type="file">
    if file and file.filename:  # new file uploaded
        filename = secure_filename(file.filename)
        file_path = RECEIPTS_DIR / filename
        file.save(file_path)
        receipt_path = filename
    else:  # no new file uploaded, keep old path
        receipt_path = current_receipt


    data = (
        request.form["date"],
        request.form.get("payment_ref", ""),
        request.form.get("factura_ref", ""),
        int(request.form["proveedor_id"]),
        request.form["payment_type"],
        float(request.form["amount"]),
        request.form["currency"],
        request.form.get("details", ""),
        delivered_email,
        factura_aparte,
        receipt_path,
        id,
    )
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE expenses
            SET date=?, payment_ref=?, factura_ref=?, proveedor_id=?, payment_type=?, amount=?, currency=?, details=?, delivered_email=?, factura_aparte=?, receipt_path=? 
            WHERE id=?
            """,
            data,
        )
    return redirect(url_for("index"))


@app.route("/delete/<int:id>", methods=["POST"])
def delete(id):
    with get_conn() as conn:
        conn.execute("DELETE FROM expenses WHERE id=?", (id,))
    return redirect(url_for("index"))


@app.route("/providers/add", methods=["POST"])
def add_provider():
    # Get the raw input from the form
    raw_names = request.form.get("name", "").strip()
    if not raw_names:
        return redirect(url_for("index"))

    names = [n.strip().title() for n in raw_names.replace("\n", ",").split(",") if n.strip()]

    if names:
        with get_conn() as conn:
            # Insert each provider, ignore duplicates
            conn.executemany("INSERT OR IGNORE INTO providers(name) VALUES (?)", [(n,) for n in names])

    return redirect(url_for("index"))




@app.route("/export", methods=["GET"])
def export_excel():
    month = request.args.get("month")
    provider = request.args.get("provider")
    payment_type = request.args.get("payment_type")

    query = """
        SELECT e.date, p.name as proveedor, e.payment_type, e.amount, e.currency,
               e.factura_ref, e.payment_ref, e.delivered_email, e.factura_aparte, e.details
        FROM expenses e
        JOIN providers p ON p.id = e.proveedor_id
        WHERE 1=1
    """
    params = []
    if month:
        query += " AND strftime('%Y-%m', e.date) = ?"
        params.append(month)
    if provider:
        query += " AND e.proveedor_id = ?"
        params.append(provider)
    if payment_type:
        query += " AND e.payment_type = ?"
        params.append(payment_type)
    query += " ORDER BY e.date DESC"

    with get_conn() as conn:
        rows = conn.execute(query, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # --- Styles ---
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    currency_style = NamedStyle(name="currency_style", number_format="#,##0.00")

    # Timestamp header
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["Fecha", "Proveedor", "Pago", "Monto", "Moneda", "Factura Ref", "Ref Pago", "Enviado Email", "Factura Aparte", "Detalles"]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        c = ws.cell(row=2, column=col)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.border = border
        if col == 4:
            c.alignment = Alignment(horizontal="right")

    # Data rows
    for r in rows:
        ws.append([
            r["date"],
            r["proveedor"],
            r["payment_type"].capitalize(),
            r["amount"],
            r["currency"],
            r["factura_ref"] or "",
            r["payment_ref"] or "",
            "Sí" if r["delivered_email"] else "No",
            "Sí" if r["factura_aparte"] else "No",
            r["details"] or ""
        ])

    # Apply styles & compute totals
    total_crc = 0
    total_usd = 0
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=10), start=3):
        amt_cell = row[3]  # "Monto"
        amt_cell.border = border
        amt_cell.alignment = Alignment(horizontal="right")
        if row[4].value == "CRC":
            total_crc += amt_cell.value
        elif row[4].value == "USD":
            total_usd += amt_cell.value
        for cell in row:
            if cell.column != 4:  # other cells
                cell.border = border

    # Add totals row
    total_row = ws.max_row + 2
    ws[f"C{total_row}"] = "Total CRC:"
    ws[f"C{total_row}"].font = Font(bold=True)
    ws[f"D{total_row}"] = total_crc
    ws[f"D{total_row}"].number_format = "#,##0.00"
    ws[f"D{total_row}"].font = Font(bold=True)
    ws[f"C{total_row+1}"] = "Total USD:"
    ws[f"C{total_row+1}"].font = Font(bold=True)
    ws[f"D{total_row+1}"] = total_usd
    ws[f"D{total_row+1}"].number_format = "#,##0.00"
    ws[f"D{total_row+1}"].font = Font(bold=True)

    # Adjust column widths
    widths = [12, 20, 12, 12, 8, 15, 15, 12, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


if __name__ == "__main__":
    # Create DB if not exists and ensure schema
    if not Path(DB).exists():
        Path(DB).touch()  # created; init_db() will structure it
    init_db()
    app.run(debug=True)
